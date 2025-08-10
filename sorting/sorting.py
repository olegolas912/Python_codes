# -*- coding: utf-8 -*-
"""
Сводит raport.xlsx + шахмат.xlsx → output.xlsx
(несколько месяцев и лет обрабатываются за один запуск).

Теперь корректно работает со скважинами, у которых в номере есть буквы
(«104А», «25-Ю» и т.п.): номер скважины хранится как строка, а лист
в шахматке ищется без учёта регистра и пробелов.
"""

from __future__ import annotations

import calendar
from pathlib import Path
from collections import defaultdict
import pandas as pd
import openpyxl

# ------------------------------------------------------------------------------
CONFIG = {
    # файлы
    "raport_file": Path("raport восточный уренгой.xlsx"),
    "shakmat_file": Path("шахмат.xlsx"),
    "output_file": Path("output.xlsx"),

    # фильтры
    "target_field": "уренгойское",
    "target_lic":   "восточно-уренгойское",

    # допустимые секции
    "allowed_sections": {
        "газоконденсатные скважины новые (введенные в текущем месяце)",
        "газоконденсатные скважины старые (введенные до текущего месяца)",
    },

    # индексы колонок raport (нумерация с 0)
    "cols": {                    #   A   B   C   D   …
        "well": 1,               #   0   1   2   3
        "gas":  (3, 4),          #   ↑   ↑
        "cond": 9,
        "water": 12,
        "days": 5,
    },

    # ярлыки строк шахматки
    "labels": {
        "pbuf": "Pбуф, атм",
        "pzat": "Pзат, атм",
        "dsht": "Dшт, мм",
    },

    # рус-месяц → номер
    "rus_months": {
        "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
        "май": 5, "июнь": 6, "июль": 7, "август": 8,
        "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
    },
}
# ------------------------------------------------------------------------------


# ---------- утилиты -----------------------------------------------------------
def days_in_month(month_ru: str, year: int) -> int:
    m = CONFIG["rus_months"][month_ru]
    return calendar.monthrange(year, m)[1]


def locate_header(sheet, month_ru: str, year: int) -> int | None:
    """
    В шахматке находим строку, где начинается блок нужного (месяц, год).

    • Строка-год:    B — целое число.
    • Строка-месяц:  A содержит <месяц>, B = 'Дата'.
    """
    cur_year = None
    month_ru = month_ru.lower()

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        a, b = row[:2]

        # строка-год
        if isinstance(b, (int, float)) and float(b).is_integer():
            cur_year = int(b)
            continue

        # строка-месяц
        if (cur_year == year and
            isinstance(a, str) and month_ru in a.lower() and
            isinstance(b, str) and b.strip().lower() == "дата"):
            return i
    return None


def col_idx(header_row: list, name: str) -> int | None:
    name = name.lower().strip()
    for j, v in enumerate(header_row):
        if isinstance(v, str) and v.lower().strip() == name:
            return j
    return None


def last_daily(sheet, month_ru: str, year: int, label: str) -> float | None:
    hi = locate_header(sheet, month_ru, year)
    if hi is None:
        return None

    rows = list(sheet.iter_rows(values_only=True))[hi:]
    hdr = rows[0]
    last_day_col = max(j for j, v in enumerate(hdr)
                       if isinstance(v, (int, float)) and float(v).is_integer())

    for r in rows[1:]:
        if r[1] == "Дата":           # следующий месяц
            break
        if str(r[1]).lower().strip() == label.lower():
            for k in range(last_day_col, 1, -1):
                if k < len(r) and isinstance(r[k], (int, float)):
                    return r[k]
    return None


def avg_value(sheet, month_ru: str, year: int, label: str) -> float | None:
    hi = locate_header(sheet, month_ru, year)
    if hi is None:
        return None

    rows = list(sheet.iter_rows(values_only=True))[hi:]
    hdr = rows[0]
    avg_col = col_idx(hdr, "среднее")
    if avg_col is None:
        return None

    for r in rows[1:]:
        if r[1] == "Дата":
            break
        if str(r[1]).lower().strip() == label.lower():
            v = r[avg_col]
            return v if isinstance(v, (int, float)) else None
    return None
# ------------------------------------------------------------------------------


def process() -> pd.DataFrame:
    out = []

    # ---------- 1. Загружаем шахматку (read-only) ----------------------------
    wb_shak = openpyxl.load_workbook(
        CONFIG["shakmat_file"], data_only=True, read_only=True
    )
    # карта «нормализованное имя листа» → «реальное имя листа»
    sheet_map = {s.lower().strip(): s for s in wb_shak.sheetnames}

    # ---------- 2. Обрабатываем каждый подходящий лист RAPORT ---------------
    wb_rap = openpyxl.load_workbook(CONFIG["raport_file"], read_only=True)

    allowed = {s.lower() for s in CONFIG["allowed_sections"]}
    cols = CONFIG["cols"]

    for sheet_name in wb_rap.sheetnames:
        parts = sheet_name.split()
        if len(parts) != 2 or parts[0].lower() not in CONFIG["rus_months"]:
            continue                       # пропускаем «левые» листы

        month_ru, year = parts[0].lower(), int(parts[1])
        n_days = days_in_month(month_ru, year)

        df = pd.read_excel(
            CONFIG["raport_file"], sheet_name=sheet_name,
            header=None, dtype=object
        )
        wells = defaultdict(lambda: dict(gas=0.0, cond=0.0,
                                         water=0.0, days=0.0, n=0))

        cur_field = cur_lic = None
        in_section = False

        for _, row in df.iterrows():
            first = row.iloc[0]

            # ── заголовочные строки ─────────────────────────────────────────
            if isinstance(first, str):
                txt = first.lower().strip()
                if txt.startswith("месторождение"):
                    cur_field = txt.split(":", 1)[1].strip(); in_section = False
                elif txt.startswith("лицензионный участок"):
                    cur_lic = txt.split(":", 1)[1].strip();   in_section = False
                else:
                    in_section = txt in allowed
                continue

            # ── строки-данные ───────────────────────────────────────────────
            if (not in_section or
                cur_field != CONFIG["target_field"] or
                cur_lic   != CONFIG["target_lic"]):
                continue

            well_raw = row.iloc[cols["well"]]
            if pd.isna(well_raw):
                continue

            # безопасное преобразование номера скважины в строку
            if isinstance(well_raw, (int, float)) and float(well_raw).is_integer():
                well = str(int(well_raw))          # 104.0 → '104'
            else:
                well = str(well_raw).strip()       # '104А' → '104А'

            gas = sum(row.iloc[c] for c in cols["gas"] if not pd.isna(row.iloc[c]))

            rec = wells[well]
            rec["gas"]   += gas
            rec["cond"]  += row.iloc[cols["cond"]]  if not pd.isna(row.iloc[cols["cond"]])  else 0
            rec["water"] += row.iloc[cols["water"]] if not pd.isna(row.iloc[cols["water"]]) else 0
            rec["days"]  += row.iloc[cols["days"]]  if not pd.isna(row.iloc[cols["days"]])  else 0
            rec["n"]     += 1

        # ---------- 3. Добавляем данные шахматки ----------------------------
        for well, p in wells.items():
            coef = p["days"] / (n_days * p["n"]) if p["n"] else None

            pbuf = pzat = dsht = None
            key = well.lower().strip()             # нормализуем имя листа
            if key in sheet_map:                   # ищем без учёта регистра
                sh = wb_shak[sheet_map[key]]
                pbuf = last_daily(sh, month_ru, year, CONFIG["labels"]["pbuf"])
                pzat = last_daily(sh, month_ru, year, CONFIG["labels"]["pzat"])
                dsht = avg_value (sh, month_ru, year, CONFIG["labels"]["dsht"])

            out.append({
                "Скважина": well,
                "Дата": f"{year}-{CONFIG['rus_months'][month_ru]:02d}-{n_days:02d}",
                "Коэф. раб.": round(coef, 3) if coef is not None else None,
                "Gas, тыс.м³": round(p["gas"], 2),
                "Cond, т":    round(p["cond"], 2),
                "Water, м³":  round(p["water"], 2),
                "Pбуф, атм":  pbuf,
                "Pзат, атм":  pzat,
                "Dшт, мм":    dsht,
            })

    return pd.DataFrame(out)
# ------------------------------------------------------------------------------


if __name__ == "__main__":
    df = process().sort_values(["Дата", "Скважина"])
    df.to_excel(CONFIG["output_file"], index=False)
    print(f"✓ Результат сохранён в {CONFIG['output_file'].resolve()}")
