import time
import openpyxl
from openpyxl import load_workbook

import re
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def login_elibrary(driver, username, password):
    driver.get("https://elibrary.ru")

    login_input = WebDriverWait(driver, 1).until(
        EC.element_to_be_clickable((By.ID, "login"))
    )
    login_input.send_keys(username)

    password_input = WebDriverWait(driver, 1).until(
        EC.element_to_be_clickable((By.ID, "password"))
    )
    password_input.send_keys(password)

    password_input.send_keys(Keys.ENTER)
    time.sleep(3)


def get_citation_text(driver):
    # Переключаемся во фрейм fancybox
    WebDriverWait(driver, 1).until(
        EC.frame_to_be_available_and_switch_to_it((By.ID, "fancybox-frame"))
    )

    # Ждем div id="ref"
    WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.ID, "ref"))
    )
    ref_div = driver.find_element(By.ID, "ref")
    citation_text = ref_div.text

    # Возвращаемся в основной документ
    driver.switch_to.default_content()
    return citation_text


def process_single_link(driver, link):
    driver.get(link)
    time.sleep(1)

    # Ссылка "Ссылка для цитирования"
    WebDriverWait(driver, 1).until(
        EC.presence_of_element_located((By.LINK_TEXT, "Ссылка для цитирования"))
    )
    driver.find_element(By.LINK_TEXT, "Ссылка для цитирования").click()

    time.sleep(1)
    citation = get_citation_text(driver)
    print(citation)
    return citation


def main():
    chrome_options = Options()
    # chrome_options.add_argument("--headless")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        username = ""
        password = ""
        login_elibrary(driver, username, password)

        excel_file = "links1.xlsx"
        wb = load_workbook(excel_file)
        sheet = wb.active

        for row in range(1, sheet.max_row + 1):
            link_cell = sheet.cell(row=row, column=1)
            link_value = link_cell.value
            if not link_value:
                continue  # пропускаем пустые ссылки

            # Проверяем столбец 3 (C). Если там уже что-то есть - пропускаем
            existing_citation_cell = sheet.cell(row=row, column=3)
            if existing_citation_cell.value:
                # Уже заполнено, пропускаем
                print(f"Строка {row}: столбец 3 уже не пуст, пропускаем.")
                continue

            # Если столбец 3 пустой, пытаемся получить цитирование
            try:
                citation_text = process_single_link(driver, link_value)
                print(f"Строка {row}, ссылка: {link_value}")
                print(f"Цитирование: {citation_text}")
            except Exception as e:
                print(f"Ошибка на строке {row}, ссылка={link_value}: {e}")
                citation_text = ""

            # Пишем в столбец 3
            sheet.cell(row=row, column=3).value = citation_text
            wb.save(excel_file)

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
